from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
list_of_headers = ['Company','Designation','Responsible Person','Emails','Tech stack','Experience','Salary','Source Link']

destination_file ="./sheets/devops.xlsx"
def get_the_header_section(headers):
    wb = Workbook()
    ws = wb.active
    ft = Font(bold=True)
    for idx,header in enumerate(headers):
        ws.cell(row= 1,column=idx+1).value = headers[idx]
        ws.cell(row=1,column=idx+1).font = ft
        ws.column_dimensions[get_column_letter(idx+1)].width = 50
    return wb

wb = get_the_header_section(list_of_headers)

#ws.append(('anjan',0,0,0,0,0,0,0))
wb.save(filename=destination_file)
wb.close()





