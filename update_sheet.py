from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

devops = './sheets/devops.xlsx'
def func():
    wb = load_workbook(devops)
    ws = wb.active
    print(list(ws.values))
    ws.append(('zoom',0,0,0,0,0,0,0))
    wb.save('./sheets/devops.xlsx')
    wb.close()
func()