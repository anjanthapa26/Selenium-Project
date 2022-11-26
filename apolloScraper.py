from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from getIndustriesTags import get_the_list_of_tags


list_of_headers = ['Company','Designation','Responsible Person','Emails','Tech stack','Experience','Salary','Source Link']
DEVOPS = './sheets/devops.xlsx'


''' commenting it for the update of the dynamic feature in input tag '''

'''
def get_the_right_company(companyName,driver):

    chooseCompany = WebDriverWait(driver, 40).until(
            EC.presence_of_element_located((By.XPATH,"//input[@placeholder='Search...']"))
    )

    chooseCompany.send_keys(companyName)
    try:
        getRightCompany = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(),'"+companyName+"')]/..//div[contains(text(),'Information technology & services')]/../../../..")))
        #getRightCompany = WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH,"//*[[contains(text(),f'{companyName}')] and [contains(text(),'Information technology & services')]]")))
        getRightCompany.click()
    except Exception as e:
        driver.refresh()
        print('Could not get the rightcompany information',e)


'''


'''Get the emails of the respective rows '''

'''
def get_the_emails(driver,elem):
    email_tab = elem.find_elements(By.TAG_NAME,'button')
    email_tab[0].click()
    print('passed')
    time.sleep(4)
    try:
        email_container = WebDriverWait(driver,20).until(EC.presence_of_element_located((By.XPATH,"//div[@class='zp_9Igty']")))
        print(email_container)
        return email_container.text

    except:
        print("There is a problem in finding the email container")
'''


def extract_info_to_excel_sheet(ws,three_details_of_company,list_of_companiesDetails):
    duplicate_company_checker_list = list(map(lambda cell:cell.value,ws[get_column_letter(2)]))[1:]
    print('Got inside extract info to excel sheet section',duplicate_company_checker_list)
    if list_of_companiesDetails[1] in duplicate_company_checker_list:
        return ws
    else:
        for rows in three_details_of_company:
            print(three_details_of_company,list_of_companiesDetails)
            ws.append([list_of_companiesDetails[0],list_of_companiesDetails[1],rows[0],rows[1],rows[2],list_of_companiesDetails[5],list_of_companiesDetails[2],list_of_companiesDetails[3],list_of_companiesDetails[4]])

        return ws


def get_the_emails(driver,elem):
    elem.click()

    try:
        if_element_is_hidden = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,"//div[normalize-space()='Access Email & Phone']/..")))
        if_element_is_hidden.click()
        if_elem_not_hidden = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,"//div[@class='zp-contact-email-envelope-container zp_n4sev zp_1sjoN']/div[1]/a")))
        return if_elem_not_hidden.text
    except:
        try:
            if_elem_not_hidden = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.XPATH,"//div[@class='zp-contact-email-envelope-container zp_n4sev zp_1sjoN']/div[1]/a")))
            return if_elem_not_hidden.text
        except:
            return 'No Email'
''' Methods to get a different tbodies '''

def get_different_tbodies(driver,get_rows):

    try:
        table_rows_datas = get_rows.find_elements(By.CLASS_NAME,'zp_1sEIg')
    except:
        print('no table_rows_datas')
    try:
        responsible_person = table_rows_datas[0].find_element(By.CLASS_NAME,value="zp_EqOJn")
        print(responsible_person.text)

    except:
        print('no responsible_person')

    try:
        Designation = table_rows_datas[1].text
        print(Designation)
    except:
        print('No designation')

    try:
        email = get_the_emails(driver,table_rows_datas[0])
    except:
        print('No email')

    return [Designation,responsible_person.text,email]

''' Get the items such as Name, designation and emails of the respective company'''

def get_the_precise_details_of_company(driver):
    counter = 1
    companies_precise_details = []
    try:
        get_tables =  WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME,"zp_lXHd4")))

        for get_rows in get_tables:
            if counter <= 3:
                companies_info = get_different_tbodies(driver,get_rows)
                companies_precise_details.append(companies_info)
            counter +=1
        return companies_precise_details
        

    except Exception as e:
        return e


''' Dynamic ways to find the list of people'''

def get_people_somehow(driver):
    plist = ['People','Employees']
    count = 0
    while count < 2:
        try:
            g_people = driver.find_element(By.XPATH,"//a[normalize-space()='"+plist[count]+"']")
            return g_people
        except:
            count +=1
            continue



''' Get the further details of a company once the company if verifeid ''' 
def get_details_of_company(driver):
    static_requirement_of_client = ['CTO','Talent acquisition','CFO','CRO']

    get_people = get_people_somehow(driver)
    get_people.click()

    ''' open the filter tab '''
    while True:
        try:
            get_filter = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,"//*[contains(text(),'Open Filters')]/.."))
            )
    
            get_filter.click()
            break
        except:
            driver.refresh()
            continue


    ''' click on the job title and enter the criterias for the jobs post on the search bar '''

    try:
        get_job = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH,"//*[contains(text(),'Job Titles')]/../.."))
        )
        get_job.click()
        time.sleep(2)
        try:
            get_input_select = driver.find_element(By.XPATH,"//div[contains(text(),'Search for a job title')]/../input")

        except:
            print('unable to get the Search for a job title')
        
        for job_titles in static_requirement_of_client:
            get_input_select.send_keys(job_titles)
            time.sleep(2)
            get_input_select.send_keys(Keys.RETURN)

        time.sleep(2)
        apply_filter = driver.find_element(By.XPATH,"//*[contains(text(),'Apply Filters')]/..")
        apply_filter.click()
        time.sleep(2)
        return get_the_precise_details_of_company(driver)

    except:
        print('found some error on job title section')

    '''Get the items such as name , designation and emails from the respective company '''


    
def check_if_revenue_matches_criteria(driver):
    try:
        checkRevenue = WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(),'Annual Revenue')]/../div[2]"))).text
        print(checkRevenue)
        if checkRevenue[-1] == 'B' or (checkRevenue[-1] == 'M' and int(checkRevenue[1:len(checkRevenue)-1]) > 50):
            return False
        else:
            return True

    except Exception as e:
        return True


''' Commenting this because implementing new features on input field to handle it dynamically '''

'''
def find_if_eligible_company(driver,list_of_companiesDetails):

    wb,ws = get_the_header_section(list_of_headers)
    for com_details in list_of_companiesDetails:

        # should fix this issue of not getting the company name 

        if com_details[0] != 'No company name available':


            get_the_right_company(com_details[0],driver)
            if check_if_revenue_matches_criteria(driver) == True:
                three_details_of_company = get_details_of_company(driver)
                if len(three_details_of_company) > 0:
                    ws = extract_info_to_excel_sheet(ws,three_details_of_company,com_details)

        break

    wb.save(filename = dest_filename)

'''


def find_if_eligible_company(driver,list_of_companiesDetails):

    wb = load_workbook(filename=DEVOPS)
    ws = wb.active
    print(list(ws.values))
    for com_details in list_of_companiesDetails:

        # should fix this issue of not getting the company name 

        total_lists_companies_of_respective_company = get_the_list_of_tags(com_details[1],driver)

        print(total_lists_companies_of_respective_company)
        if not total_lists_companies_of_respective_company:
            driver.refresh()
            continue
        counter = 0
        while counter < len(total_lists_companies_of_respective_company):
            print('inside while loop', counter,total_lists_companies_of_respective_company[counter][0],com_details[1])
            if total_lists_companies_of_respective_company[counter][0].lower() == com_details[1].lower():

                print(total_lists_companies_of_respective_company[counter])
                try:
                    getRightCompany = WebDriverWait(driver,7).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(),'"+com_details[1]+"')]/..//div[contains(text(),'"+total_lists_companies_of_respective_company[counter][1]+"')]/../../../..")))
                    #exact_xpath ="//*[@class='zp_1RaZe zp_1omjm'][contains(normalize-space(),'"+com_details[0]+"')][contains(normalize-space(),'"+total_lists_companies_of_respective_company[counter][1]+"')]"
                    #getRightCompany = WebDriverWait(driver,30).until(EC.element_to_be_clickable(("//*[@class='zp_2brNs'][contains(normalize-space(),'"+com_details[0]+"')][contains(normalize-space(),'"+total_lists_companies_of_respective_company[counter][1]+"')]/div[1]")))
                    #getRightCompany = WebDriverWait(driver,30).until(EC.presence_of_element_located((By.XPATH,"//*[contains(text(),'"+com_details[0]+"')]/..//div[contains(text(),'"+total_lists_companies_of_respective_company[counter][1]+"')]/../../..")))
                    getRightCompany.click()
                except:
                    print('Unable to fetch the tag data')

                    ''' Getting the perviously used code '''

                try:
                    if check_if_revenue_matches_criteria(driver) == True:
                        three_details_of_company = get_details_of_company(driver)


                        if len(three_details_of_company) > 0:
                            ws = extract_info_to_excel_sheet(ws,three_details_of_company,com_details)


                    if counter != len(total_lists_companies_of_respective_company)-1:
                        get_to_chooseCompany = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,"//input[@placeholder='Search...']")))
                        get_to_chooseCompany.clear()
                        get_to_chooseCompany.send_keys(com_details[1].split(',')[0])


                except Exception as e:
                    print('The company does not contains details according to the given tags')
                finally:
                    counter += 1

            else:
                counter += 1
                continue
        else:
            driver.refresh()

    wb.save(filename = DEVOPS)

    wb.close()


            













